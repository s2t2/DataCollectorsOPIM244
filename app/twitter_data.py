
import tweepy
from openpyxl import Workbook, load_workbook

import datetime


def twitterAPIfunction(TWITTER_ACCESS_TOKEN_para, TWITTER_ACCESS_TOKEN_SECRET_para, TWITTER_API_KEY_para, 
                                    TWITTER_API_KEY_SECRET_para, user_API_type_para, user_twitter_para):

    twitterItem = user_twitter_para
    numberOfTweets = 500

    auth = tweepy.OAuthHandler(TWITTER_API_KEY_para, TWITTER_API_KEY_SECRET_para)
    auth.set_access_token(TWITTER_ACCESS_TOKEN_para, TWITTER_ACCESS_TOKEN_SECRET_para)

    api = tweepy.API(auth)


    if user_API_type_para == "account":
        #create the cursor that will find Twitter accounts
        cursor = tweepy.Cursor(api.user_timeline, id = twitterItem, tweet_mode = "extended", wait_on_rate_limit = True).items(numberOfTweets)
                
    elif user_API_type_para == "hashtag" or user_API_type_para == "keyword":
        #create the cursor that will find Twitter accounts
        cursor = tweepy.Cursor(api.search_tweets, q = twitterItem, result_type = "popular", tweet_mode = "extended", wait_on_rate_limit = True).items(numberOfTweets)

    #create tweet class which is used to store data for each tweet
    class Tweet:
        def __init__(self, dateParameter, likesParameter, retweetsParameter, commentsParameter, textParameter):
            self.dateUnedited = dateParameter
            self.date = str(dateParameter).split()[0]
            self.timeUnedited = str(dateParameter).split()[1]
            self.likes = likesParameter
            self.retweets = retweetsParameter
            self.comments = commentsParameter
            self.text = textParameter
            
        def getDate(self):
            return self.date
        
        def getTime(self):
            return self.timeUnedited.split("+")[0]
        
        def getLikes(self):
            return self.likes
        
        def getRetweets(self):
            return self.retweets
        
        def getComments(self):
            return self.comments
        
        def getText(self):
            return self.text


    tweetsList = []

    #collect tweet data
    for i in cursor:
        tweetsList.append(Tweet(i.created_at, i.favorite_count, i.retweet_count, "None", i.full_text))
    
    #function used to print tweet data
    def printTweetData(tweetParameter):
        print("Date: " + str(tweetParameter.getDate()) + "\n" + "Time: " + str(tweetParameter.getTime()) + "\n" + 
            "Likes: " + str(tweetParameter.getLikes()) + "\n" + 
            "Retweets: " + str(tweetParameter.getRetweets()) + "\n" + "Text: " + str(tweetParameter.getText()) + "\n\n")


    #turn a tweet's text into a list and remove grammer and non-letter/non-numbers
    #splits a tweet's text into a list of words and a list of hashtags
    def prepareTweetText(tweetListNumber):
        
        tweetTextString = str(tweetsList[tweetListNumber].getText())
        newTweetTextString = ""

        #remove non-letters and non-numbers
        for character in tweetTextString:
            if character.isalnum() or character == " " or character == "#" or character == "'" or character == "-":
                newTweetTextString += character.lower()
        
        tweetTextList = newTweetTextString.split()
        tweetWordList = []
        tweetHashtagList = []
        
        for i in range(len(tweetTextList)):
            if tweetTextList[i][0] == "#":
                tweetWordList.append("")
                tweetHashtagList.append(tweetTextList[i])
            else:
                tweetWordList.append(tweetTextList[i])
                tweetHashtagList.append("")
        
        return tweetWordList, tweetHashtagList
            

    #filter out Retweets
    NumberDeleted = 0
    for i in range(len(tweetsList)):
        TestRT = str(tweetsList[i - NumberDeleted].getText())
        
        if TestRT[0:2] == "RT":
            del tweetsList[i - NumberDeleted]
            NumberDeleted += 1
            
            
    #find the weekday based on a date
    def findWeekday(dateParameter):
        
        weekdayStrings = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")

        dateList = dateParameter.split("-")

        date = datetime.date(int(dateList[0]),int(dateList[1]),int(dateList[2]))
        weekdayInteger = date.weekday()
        
        return weekdayStrings[weekdayInteger]

    #prints all Tweet data
    #WARNING: this has caused errors when there are non UTF-8 characters trying to be printed
    """
    for i in range(len(tweetsList)):
        printTweetData(tweetsList[i])
    """
                                
                        
                        
    #create output Excel file
    file_name = "web_app/" + twitterItem + " Twitter Data.xlsx"
    excelFile = Workbook()
    excelFile.save(file_name)
        
    excelFile = load_workbook(file_name)
    sheet = excelFile.active

    #sheet name
    sheet.title = "Posts"

    #add headers
    sheet.cell(row = 1, column = 1).value = "Id" 
    sheet.cell(row = 1, column = 2).value = "Date" 
    sheet.cell(row = 1, column = 3).value = "Time" 
    sheet.cell(row = 1, column = 4).value = "Weekday" 
    sheet.cell(row = 1, column = 5).value = "Likes" 
    sheet.cell(row = 1, column = 6).value = "Retweets"
    sheet.cell(row = 1, column = 7).value = "Text"

    excelFile.save(file_name)

    #adjust column widths
    sheet.column_dimensions['A'].width = 7.11
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 10.22
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 9.78
    sheet.column_dimensions['F'].width = 9.78
    sheet.column_dimensions['G'].width = 250

    excelFile.save(file_name)

    #print post data to the Excel file
    lastRowFree = 2

    for i in range(len(tweetsList)):
        sheet.cell(row = lastRowFree, column = 1).value = i + 1
        sheet.cell(row = lastRowFree, column = 2).value = tweetsList[i].getDate()
        sheet.cell(row = lastRowFree, column = 3).value = tweetsList[i].getTime()
        sheet.cell(row = lastRowFree, column = 4).value = findWeekday(str(tweetsList[i].getDate()))
        sheet.cell(row = lastRowFree, column = 5).value = tweetsList[i].getLikes()
        sheet.cell(row = lastRowFree, column = 6).value = tweetsList[i].getRetweets()
        sheet.cell(row = lastRowFree, column = 7).value = tweetsList[i].getText()
        
        lastRowFree += 1
        
    excelFile.save(file_name)

    #print the words to the Excel file
    sheet2 = excelFile.create_sheet(title="Words")
    sheet2.cell(row = 1, column = 1).value = "Post_Id" 
    sheet2.cell(row = 1, column = 2).value = "Words" 
    sheet2.column_dimensions['A'].width = 7.8
    sheet2.column_dimensions['B'].width = 21.22

    lastRowFree = 2

    for i in range(len(tweetsList)):
        listOfWords, listOfHashtags = prepareTweetText(i)
        for x in range(len(listOfWords)):
            sheet2.cell(row = lastRowFree, column = 1).value = i + 1
            sheet2.cell(row = lastRowFree, column = 2).value = listOfWords[x]
            lastRowFree += 1

    excelFile.save(file_name)

    #print the hashtags to the Excel file
    sheet3 = excelFile.create_sheet(title="Hashtags")
    sheet3.cell(row = 1, column = 1).value = "Post_Id" 
    sheet3.cell(row = 1, column = 2).value = "Hashtags" 
    sheet3.column_dimensions['A'].width = 7.8
    sheet3.column_dimensions['B'].width = 21.22

    lastRowFree = 2

    for i in range(len(tweetsList)):
        listOfWords, listOfHashtags = prepareTweetText(i)
        for x in range(len(listOfWords)):
            sheet3.cell(row = lastRowFree, column = 1).value = i + 1
            sheet3.cell(row = lastRowFree, column = 2).value = listOfHashtags[x]
            lastRowFree += 1

    excelFile.save(file_name)

    return file_name, tweetsList